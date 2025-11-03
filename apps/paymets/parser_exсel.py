import os
import xlrd
from zipfile import BadZipFile
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import xml.etree.ElementTree as ET

class ExcelPaymentParser:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.sheet = self._ensure_xlsx()

    def _ensure_xlsx(self) -> Worksheet:
        ext = os.path.splitext(self.file_path)[1].lower()

        if ext == ".xlsx":
            try:
                wb = load_workbook(self.file_path, data_only=True)
                return wb.active
            except BadZipFile:
                pass  # Proceed to try as .xls or XML

        # Check if it's XML format before trying xlrd
        with open(self.file_path, 'rb') as f:
            header = f.read(20)  # Read enough to check XML declaration

        if header.startswith(b'<?xml'):
            # Parse as Excel XML Spreadsheet
            ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
            try:
                tree = ET.parse(self.file_path)
                root = tree.getroot()
                worksheet = root.find('ss:Worksheet', ns)
                if worksheet is None:
                    raise ValueError("No Worksheet found in XML")
                table = worksheet.find('ss:Table', ns)
                if table is None:
                    raise ValueError("No Table found in XML")

                wb_xlsx = Workbook()
                ws_xlsx = wb_xlsx.active

                row_idx = 1
                for row in table.findall('ss:Row', ns):
                    col_idx = 1
                    for cell in row.findall('ss:Cell', ns):
                        data = cell.find('ss:Data', ns)
                        if data is not None:
                            val = data.text
                            typ = data.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}Type')
                            if typ == 'Number':
                                try:
                                    val = float(val)
                                except ValueError:
                                    pass  # Keep as string if conversion fails
                            # Add handling for other types like DateTime if needed
                            ws_xlsx.cell(row=row_idx, column=col_idx, value=val)
                        col_idx += 1
                    row_idx += 1
                return ws_xlsx
            except ET.ParseError:
                raise ValueError("Invalid XML structure in file")

        # Fallback to xlrd for binary .xls
        try:
            book_xls = xlrd.open_workbook(self.file_path)
            sheet_xls = book_xls.sheet_by_index(0)
        except xlrd.XLRDError:
            raise ValueError("Файл не является валидным .xls/.xlsx/XML, не удалось открыть")

        wb_xlsx = Workbook()
        ws_xlsx = wb_xlsx.active

        for r in range(sheet_xls.nrows):
            for c in range(sheet_xls.ncols):
                ws_xlsx.cell(row=r + 1, column=c + 1).value = sheet_xls.cell_value(r, c)
        return ws_xlsx

    def extract_kazpost_data(self):
        """
        Извлекает данные из листа openpyxl (ТОЛЬКО xlsx-совместимый sheet).
        Возвращает список словарей: № лицевого счета, дата, сумма платежа, № операции.
        """

        sheet = self.sheet
        # 1. Ищем строку заголовка по "№ п/п"
        header_row_idx = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row and row[0] and "№ п/п" in str(row[0]):
                header_row_idx = idx
                break
        if not header_row_idx:
            return []

        # 2. Находим индексы нужных столбцов
        headers_cells = list(sheet[header_row_idx])
        account_col = None
        amount_col = None
        operation_col = None

        for i, cell in enumerate(headers_cells, start=1):
            val = str(cell.value) if cell.value is not None else ""
            if "лицевого счета" in val:
                account_col = i
            if ("Сумма оплаты" in val) or (val == "Сумма"):
                amount_col = i
            if "№ операции" in val or "номер операции" in val.lower():
                operation_col = i

        # 3. Извлекаем дату из заголовка документа
        document_date = None
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and row[0] and "на дату:" in str(row[0]):
                document_date = str(row[0]).replace("на дату:", "").strip()
                break

        # 4. Собираем данные
        results = []
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row:
                continue

            account = row[account_col - 1] if account_col and len(row) >= account_col else None
            amount = row[amount_col - 1] if amount_col and len(row) >= amount_col else None
            operation = row[operation_col - 1] if operation_col and len(row) >= operation_col else None

            # Пропускаем строки без счета/суммы или с "Итого"
            if not account or not amount or "Итого" in str(amount):
                continue

            # Форматируем лицевой счет
            acc_str = str(int(account)) if isinstance(account, float) else str(account)

            # Форматируем сумму
            try:
                amount_f = float(amount)
            except:
                amount_f = 0.0

            # Форматируем № операции
            op_str = ""
            if operation is not None:
                op_str = str(int(operation)) if isinstance(operation, float) else str(operation)

            results.append({
                "Date": document_date,
                "Account": acc_str,
                "PaymentID": op_str,
                "Amount": amount_f
            })

        print(results)
        return results

    def extract_kaspi_data(self):
        """
        Парсит отчет Kaspi (лист 'Данные') и возвращает список словарей:
        - 'Дата' (строка как в файле)
        - 'Идентификатор платежа' (строка)
        - 'Лицевой счет' (строка)
        - 'Сумма платежа' (float)
        """

        sheet = self.sheet
        # Найти строку заголовков колонок
        header_row_idx = None
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row:
                continue
            # Ищем шапку с явными названиями
            if str(row[0]).strip() == "Дата" and str(row[2]).strip() == "Лицевой счет":
                header_row_idx = i
                break

        if not header_row_idx:
            return []

        # Индексы колонок по ожидаемым именам
        header_cells = list(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
        col_map = {}
        for idx, name in enumerate(header_cells):
            if name is None:
                continue
            name_str = str(name).strip()
            if name_str == "Дата":
                col_map["date"] = idx
            elif name_str == "Идентификатор платежа":
                col_map["payment_id"] = idx
            elif name_str == "Лицевой счет":
                col_map["account"] = idx
            elif name_str == "Сумма платежа":
                col_map["amount"] = idx

        # Проверка, что все нужные колонки найдены
        if not {"date", "payment_id", "account", "amount"}.issubset(col_map.keys()):
            return []

        results = []
        # Данные начинаются со следующей строки
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row:
                continue
            # Остановка при встрече блока итогов
            first_cell = str(row[0]).strip() if row[0] is not None else ""
            if first_cell in {"Общая сумма", "Комиссия", "Сумма к перечислению", "Количество"}:
                break

            date_val = row[col_map["date"]]
            payment_id_val = row[col_map["payment_id"]]
            acc_val = row[col_map["account"]]
            amt_val = row[col_map["amount"]]

            # Пропускаем пустые строки
            if acc_val in (None, "") or amt_val in (None, "") or date_val in (None, ""):
                continue

            # Приведение типов
            # Дату оставим как строку (как в файле), но если это datetime — превратим в YYYY-MM-DD
            if hasattr(date_val, "strftime"):
                date_out = date_val.strftime("%Y-%m-%d")
            else:
                date_out = str(date_val)

            # Идентификатор платежа
            payment_id_out = ""
            if payment_id_val is not None:
                payment_id_out = str(int(payment_id_val)) if isinstance(payment_id_val, (int, float)) else str(payment_id_val)

            account_out = str(int(acc_val)) if isinstance(acc_val, (int, float)) else str(acc_val)

            try:
                amount_out = float(amt_val)
            except:
                continue

            results.append({
                "Date": date_out,
                "Account": account_out,
                "PaymentID": payment_id_out,
                "Amount": amount_out
            })
        print(results)
        return results

    def extract_halyk_data(self):
        """
        Парсит отчет Halyk и возвращает список словарей:
        - 'Дата' (строка)
        - 'Идентификатор платежа' (строка)
        - 'Лицевой счет' (строка)
        - 'Сумма платежа' (float)
        """

        sheet = self.sheet
        # Найти строку заголовков колонок
        header_row_idx = None
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row:
                continue
            # Ищем шапку по ключевым словам
            first_cell = str(row[0]).strip() if row[0] else ""
            if "Дата операционного дня" in first_cell or first_cell == "Дата операционного дня":
                header_row_idx = i
                break

        if not header_row_idx:
            return []

        # Индексы колонок по ожидаемым именам
        header_cells = list(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
        col_map = {}
        for idx, name in enumerate(header_cells):
            if name is None:
                continue
            name_str = str(name).strip()
            if "Дата операционного дня" in name_str:
                col_map["date"] = idx
            elif "Идентификатор платежа" in name_str:
                col_map["payment_id"] = idx
            elif "Лицевой счет" in name_str:
                col_map["account"] = idx
            elif "Сумма платежа" in name_str:
                col_map["amount"] = idx

        # Проверка, что все нужные колонки найдены
        if not {"date", "payment_id", "account", "amount"}.issubset(col_map.keys()):
            return []

        results = []
        # Данные начинаются со следующей строки
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row:
                continue

            # Остановка при встрече блока итогов
            first_cell = str(row[0]).strip() if row[0] is not None else ""
            if first_cell in {"Общая сумма", "Комиссия", "Сумма к перечислению", "Количество"}:
                break

            date_val = row[col_map["date"]] if len(row) > col_map["date"] else None
            payment_id_val = row[col_map["payment_id"]] if len(row) > col_map["payment_id"] else None
            acc_val = row[col_map["account"]] if len(row) > col_map["account"] else None
            amt_val = row[col_map["amount"]] if len(row) > col_map["amount"] else None

            # Пропускаем пустые строки
            if acc_val in (None, "") or amt_val in (None, "") or date_val in (None, ""):
                continue

            # Приведение типов
            # Дата: если datetime — в YYYY-MM-DD, иначе как строка
            if hasattr(date_val, "strftime"):
                date_out = date_val.strftime("%Y-%m-%d")
            else:
                # Формат DD/MM/YYYY -> YYYY-MM-DD
                date_str = str(date_val).strip()
                if "/" in date_str and len(date_str) == 10:
                    parts = date_str.split("/")
                    if len(parts) == 3:
                        date_out = f"{parts[2]}-{parts[1]}-{parts[0]}"
                    else:
                        date_out = date_str
                else:
                    date_out = date_str

            # Идентификатор платежа
            payment_id_out = ""
            if payment_id_val is not None:
                payment_id_out = str(int(payment_id_val)) if isinstance(payment_id_val, (int, float)) else str(payment_id_val)

            # Лицевой счет
            account_out = str(int(acc_val)) if isinstance(acc_val, (int, float)) else str(acc_val)

            # Сумма платежа
            try:
                amount_out = float(amt_val)
            except:
                continue

            results.append({
                "Date": date_out,
                "Account": account_out,
                "PaymentID": payment_id_out,
                "Amount": amount_out
            })
        print(results)
        return results

    def extract_bcc_data(self):
        sheet = self.sheet

        # Найти строку заголовков колонок
        header_row_idx = None
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row:
                continue
            # Ищем шапку по первым колонкам
            if (str(row[0]).strip() == "№" and
                    row[1] and "Плательщик" in str(row[1]) and
                    row[2] and "Дата" in str(row[2])):
                header_row_idx = i
                break

        if not header_row_idx:
            return []

        # Индексы колонок по ожидаемым именам
        header_cells = list(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
        col_map = {}
        for idx, name in enumerate(header_cells):
            if name is None:
                continue
            name_str = str(name).strip()
            if name_str == "Дата":
                col_map["date"] = idx
            elif "№ платежа" in name_str:
                col_map["payment_id"] = idx
            elif "Лицевой счет" in name_str:
                col_map["account"] = idx
            elif name_str == "Сумма":
                col_map["amount"] = idx

        # Проверка, что все нужные колонки найдены
        if not {"date", "payment_id", "account", "amount"}.issubset(col_map.keys()):
            return []

        results = []
        # Данные начинаются со следующей строки
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row:
                continue

            # Остановка при встрече строки "ИТОГО:"
            first_cell = str(row[0]).strip() if row[0] is not None else ""
            if "ИТОГО" in first_cell.upper():
                break

            date_val = row[col_map["date"]] if len(row) > col_map["date"] else None
            payment_id_val = row[col_map["payment_id"]] if len(row) > col_map["payment_id"] else None
            acc_val = row[col_map["account"]] if len(row) > col_map["account"] else None
            amt_val = row[col_map["amount"]] if len(row) > col_map["amount"] else None

            # Пропускаем пустые строки
            if acc_val in (None, "") or amt_val in (None, "") or date_val in (None, ""):
                continue

            # Приведение типов
            # Дата: если datetime — в YYYY-MM-DD, иначе как строка
            if hasattr(date_val, "strftime"):
                date_out = date_val.strftime("%Y-%m-%d")
            else:
                # Формат DD.MM.YYYY -> YYYY-MM-DD
                date_str = str(date_val).strip()
                if "." in date_str and len(date_str) == 10:
                    parts = date_str.split(".")
                    if len(parts) == 3:
                        date_out = f"{parts[2]}-{parts[1]}-{parts[0]}"
                    else:
                        date_out = date_str
                else:
                    date_out = date_str

            # Номер платежа
            payment_id_out = ""
            if payment_id_val is not None:
                payment_id_out = str(int(payment_id_val)) if isinstance(payment_id_val, (int, float)) else str(payment_id_val)

            # Лицевой счет
            account_out = str(int(acc_val)) if isinstance(acc_val, (int, float)) else str(acc_val)

            # Сумма платежа (убираем пробелы из формата "30 000.00")
            try:
                if isinstance(amt_val, str):
                    amt_cleaned = amt_val.replace(" ", "").replace("\xa0", "")
                    amount_out = float(amt_cleaned)
                else:
                    amount_out = float(amt_val)
            except:
                continue

            results.append({
                "Date": date_out,
                "Account": account_out,
                "PaymentID": payment_id_out,
                "Amount": amount_out
            })
        print(results)
        return results

